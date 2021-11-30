# -*- encoding:utf-8 -*-

import datetime as dt

import openpyxl


def stat(xlsx, start_date, end_date, accounts):
    workbook = openpyxl.load_workbook(xlsx)

    count = {}
    sum_zy = {}
    sum_yz = {}
    # for booksheet in workbook.sheets():
    #     for row in range(booksheet.nrows):
    for booksheet in workbook.worksheets:
        for row in booksheet.rows:
            # date = booksheet.cell(row, 0).value
            date = row[0].value
            if not date or date.find('-') < 0:
                continue

            # biz = booksheet.cell(row, 1).value
            biz = row[1].value
            if biz.find('银证') < 0:
                continue

            date = dt.datetime.strptime(date.strip(), '%Y-%m-%d').date()
            if date < start_date or date > end_date:
                continue

            # date = booksheet.cell(row, 0).value.strip()
            # account = booksheet.cell(row, 12).value.strip()
            # zy = booksheet.cell(row, 8).value.strip()
            # yz = booksheet.cell(row, 9).value.strip()
            date = row[0].value.strip()
            account = row[12].value.strip()
            account = account[:account.index('证券')+2]
            zy = row[8].value
            zy = zy.strip() if zy else ''
            yz = row[9].value
            yz = yz.strip() if yz else ''

            if account[:4] not in accounts:
                continue

            if account not in sum_zy:
                sum_zy[account] = [0, 0]
            if account not in sum_yz:
                sum_yz[account] = [0, 0]
            if account not in count:
                count[account] = 0

            if len(zy) > 0:  # 无值时为str, 有值时为float
                sum_zy[account][0] += float(zy.replace(',', ''))
            if len(yz) > 0:
                sum_zy[account][1] += float(yz.replace(',', ''))

            count[account] += 1

    return {'sum_zy': sum_zy, 'count': count}


def format(stat_ret, zc):
    sum_zy = stat_ret['sum_zy']
    count = stat_ret['count']
    all_count = 0
    all_zy = [0, 0]

    r = ''
    for k, v in sum_zy.items():
        _zc = 0
        all_count += count[k]
        if len(k) == 0:
            continue
        if k.find('中信证券') >= 0:
            # sum_zy[k][0] += sum_zy[''][0]
            # sum_zy[k][1] += sum_zy[''][1]
            # count[k] += count['']
            _zc = zc
        all_zy[0] += sum_zy[k][0]
        all_zy[1] += sum_zy[k][1]

        r += '{}\n次数:{}\n银证:{}\n证银:{}\n投入:{}\n盈利:{}'.format(k, count[k], sum_zy[k][0], sum_zy[k][1], round(sum_zy[k][1] - sum_zy[k][0], 2), round(_zc + sum_zy[k][0] - sum_zy[k][1], 2))
        r += '\n'

    r += '\n'
    r += '{}\n次数:{}\n银证:{}\n证银:{}\n投入:{}\n盈利:{}'.format('总计', all_count, round(all_zy[0], 2), round(all_zy[1], 2), round(all_zy[1] - all_zy[0], 2), round(zc + all_zy[0] - all_zy[1], 2))
    return r


if __name__ == '__main__':
    xlsx = '../../inv/stat/icbc20130706-20211130.xlsx'
    start_date = dt.date(2014, 10, 21)
    end_date = dt.date(2021, 11, 30)

    zc = 425837.02
    accounts = '中信证券'

    stat_ret = stat(xlsx, start_date, end_date, accounts)
    r = format(stat_ret, zc)
    print(r)
